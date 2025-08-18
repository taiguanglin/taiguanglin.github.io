#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
問答分類自動化系統
用於將舊目錄的問答內容分類到新目錄體系中
"""

import configparser
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import openai
import re
import time
import logging
from datetime import datetime
import os
from typing import Dict, List, Tuple, Optional

# 設置日誌
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(f'classification_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class QAClassifier:
    def __init__(self, config_file: str = 'config.ini'):
        """初始化分類器"""
        self.config = configparser.ConfigParser()
        self.config.read(config_file, encoding='utf-8')
        
        # 初始化OpenAI
        self.setup_openai()
        
        # 載入分類體系
        self.category_system = self.load_category_system()
        
        # 載入prompt模板
        self.prompt_template = self.load_prompt_template()
        
        logger.info("QA分類器初始化完成")
    
    def setup_openai(self):
        """設置OpenAI API"""
        api_key = self.config.get('openai', 'api_key')
        if api_key == 'YOUR_OPENAI_API_KEY_HERE':
            raise ValueError("請在config.ini中設置您的OpenAI API Key")
        
        openai.api_key = api_key
        self.model = self.config.get('openai', 'model', fallback='gpt-4')
        self.temperature = self.config.getfloat('openai', 'temperature', fallback=0.3)
        self.max_tokens = self.config.getint('openai', 'max_tokens', fallback=1000)
        
        logger.info(f"OpenAI設置完成 - 模型: {self.model}")
    
    def load_category_system(self) -> str:
        """載入新目錄分類體系"""
        toc_file = "/Users/paul/taiguanglin.github.io/new_toc/wenda2-toc-2025-08-18v2.txt"
        
        if not os.path.exists(toc_file):
            logger.error(f"找不到目錄文件: {toc_file}")
            return "目錄文件未找到，請檢查路徑"
        
        with open(toc_file, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 簡化目錄結構，提取主要分類
        lines = content.strip().split('\n')
        categories = []
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
                
            # 提取不同層級的分類
            if line.startswith('I.') or line.startswith('II.') or line.startswith('III.'):
                # 主要分類
                category = re.sub(r'\s+\(\d+\)$', '', line)
                categories.append(category)
            elif line.startswith('  ') and ('I.I.' in line or 'II.I.' in line):
                # 二級分類
                category = re.sub(r'\s+\(\d+\)$', '', line.strip())
                categories.append(category)
        
        return '\n'.join(categories[:50])  # 限制分類數量避免prompt過長
    
    def load_prompt_template(self) -> str:
        """載入prompt模板"""
        with open('prompt_template.txt', 'r', encoding='utf-8') as f:
            return f.read()
    
    def load_excel_data(self) -> Tuple[openpyxl.Workbook, openpyxl.worksheet.worksheet.Worksheet]:
        """載入Excel數據"""
        file_path = self.config.get('excel', 'file_path')
        sheet_name = self.config.get('excel', 'sheet_name')
        
        try:
            workbook = load_workbook(file_path)
            if sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
            else:
                # 如果指定的工作表不存在，使用第一個工作表
                worksheet = workbook.active
                logger.warning(f"工作表 '{sheet_name}' 不存在，使用 '{worksheet.title}'")
            
            logger.info(f"成功載入Excel文件: {file_path}, 工作表: {worksheet.title}")
            return workbook, worksheet
            
        except Exception as e:
            logger.error(f"載入Excel文件失败: {e}")
            raise
    
    def extract_qa_content(self, worksheet, row: int) -> Tuple[str, str]:
        """提取指定行的問答內容"""
        title_col = self.config.getint('excel', 'title_column')
        qa_start_col = self.config.getint('excel', 'qa_start_column')
        
        # 提取標題
        title_cell = worksheet.cell(row=row, column=title_col)
        title = str(title_cell.value) if title_cell.value else ""
        
        # 提取問答內容（從第7列開始到有內容的最後一列）
        qa_content = []
        col = qa_start_col
        max_col = worksheet.max_column
        
        while col <= max_col:
            cell = worksheet.cell(row=row, column=col)
            if cell.value:
                qa_content.append(str(cell.value))
            col += 1
        
        content = "\n\n".join(qa_content)
        return title, content
    
    def classify_qa(self, title: str, content: str) -> Dict[str, str]:
        """使用OpenAI對問答進行分類"""
        if not title and not content:
            return {
                'classification': '無法分類',
                'reason': '標題和內容均為空',
                'question_summary': '',
                'answer_summary': ''
            }
        
        # 構建prompt
        prompt = self.prompt_template.format(
            category_system=self.category_system,
            title=title,
            qa_content=content[:3000]  # 限制內容長度
        )
        
        try:
            response = openai.ChatCompletion.create(
                model=self.model,
                messages=[
                    {"role": "system", "content": "你是一個專業的佛學問答分類專家。"},
                    {"role": "user", "content": prompt}
                ],
                temperature=self.temperature,
                max_tokens=self.max_tokens
            )
            
            result_text = response.choices[0].message.content.strip()
            return self.parse_classification_result(result_text)
            
        except Exception as e:
            logger.error(f"OpenAI API調用失敗: {e}")
            return {
                'classification': 'API錯誤',
                'reason': f'API調用失敗: {str(e)}',
                'question_summary': '',
                'answer_summary': ''
            }
    
    def parse_classification_result(self, result_text: str) -> Dict[str, str]:
        """解析OpenAI的分類結果"""
        result = {
            'classification': '',
            'reason': '',
            'question_summary': '',
            'answer_summary': ''
        }
        
        try:
            # 使用正則表達式提取各部分內容
            classification_match = re.search(r'✅ 最佳分類排序：\s*\n([\s\S]*?)(?=✅|$)', result_text)
            if classification_match:
                result['classification'] = classification_match.group(1).strip()
            
            reason_match = re.search(r'✅ 理由：\s*\n([\s\S]*?)(?=✅|$)', result_text)
            if reason_match:
                result['reason'] = reason_match.group(1).strip()
            
            question_match = re.search(r'✅ 提問重點摘要：\s*\n([\s\S]*?)(?=✅|$)', result_text)
            if question_match:
                result['question_summary'] = question_match.group(1).strip()
            
            answer_match = re.search(r'✅ 回答重點摘要：\s*\n([\s\S]*?)(?=✅|$)', result_text)
            if answer_match:
                result['answer_summary'] = answer_match.group(1).strip()
                
        except Exception as e:
            logger.error(f"解析分類結果失敗: {e}")
            result['reason'] = f"解析錯誤: {str(e)}"
        
        return result
    
    def write_classification_result(self, worksheet, row: int, result: Dict[str, str]):
        """將分類結果寫入Excel"""
        classification_col = self.config.getint('output', 'classification_column')
        reason_col = self.config.getint('output', 'reason_column')
        question_summary_col = self.config.getint('output', 'question_summary_column')
        answer_summary_col = self.config.getint('output', 'answer_summary_column')
        
        # 寫入結果
        worksheet.cell(row=row, column=classification_col, value=result['classification'])
        worksheet.cell(row=row, column=reason_col, value=result['reason'])
        worksheet.cell(row=row, column=question_summary_col, value=result['question_summary'])
        worksheet.cell(row=row, column=answer_summary_col, value=result['answer_summary'])
    
    def process_batch(self, start_row: int = None, end_row: int = None):
        """批量處理問答分類"""
        # 載入配置
        if start_row is None:
            start_row = self.config.getint('processing', 'start_row', fallback=2)
        if end_row is None:
            config_end_row = self.config.getint('processing', 'end_row', fallback=0)
            end_row = config_end_row if config_end_row > 0 else None
        
        # 載入Excel
        workbook, worksheet = self.load_excel_data()
        
        # 確定處理範圍
        max_row = worksheet.max_row
        if end_row is None or end_row > max_row:
            end_row = max_row
        
        logger.info(f"開始處理第 {start_row} 到 {end_row} 行，共 {end_row - start_row + 1} 條記錄")
        
        processed_count = 0
        success_count = 0
        
        for row in range(start_row, end_row + 1):
            try:
                # 檢查是否已有分類結果
                classification_col = self.config.getint('output', 'classification_column')
                existing_classification = worksheet.cell(row=row, column=classification_col).value
                
                if existing_classification:
                    logger.info(f"第 {row} 行已有分類結果，跳過")
                    continue
                
                # 提取問答內容
                title, content = self.extract_qa_content(worksheet, row)
                
                if not title and not content:
                    logger.info(f"第 {row} 行無內容，跳過")
                    continue
                
                logger.info(f"處理第 {row} 行: {title[:50]}...")
                
                # 進行分類
                result = self.classify_qa(title, content)
                
                # 寫入結果
                self.write_classification_result(worksheet, row, result)
                
                processed_count += 1
                if result['classification'] != 'API錯誤':
                    success_count += 1
                
                logger.info(f"第 {row} 行處理完成")
                
                # API限制：適當延遲
                time.sleep(1)
                
                # 每10條記錄保存一次
                if processed_count % 10 == 0:
                    self.save_workbook(workbook)
                    logger.info(f"已處理 {processed_count} 條記錄，中間保存完成")
                
            except Exception as e:
                logger.error(f"處理第 {row} 行時發生錯誤: {e}")
                continue
        
        # 最終保存
        self.save_workbook(workbook)
        
        logger.info(f"批量處理完成！總共處理 {processed_count} 條記錄，成功 {success_count} 條")
    
    def save_workbook(self, workbook):
        """保存工作簿"""
        try:
            output_file = f"classified_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            workbook.save(output_file)
            logger.info(f"工作簿已保存為: {output_file}")
        except Exception as e:
            logger.error(f"保存工作簿失敗: {e}")

def main():
    """主函數"""
    print("問答分類自動化系統")
    print("=" * 50)
    
    try:
        classifier = QAClassifier()
        
        # 可以指定處理範圍，例如：
        # classifier.process_batch(start_row=2, end_row=10)  # 只處理前幾行測試
        
        # 處理所有記錄
        classifier.process_batch()
        
    except Exception as e:
        logger.error(f"程序執行失敗: {e}")
        print(f"錯誤: {e}")

if __name__ == "__main__":
    main()
